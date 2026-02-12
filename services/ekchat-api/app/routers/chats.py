"""Chat, web search, and data endpoints for ekchat-api."""
from __future__ import annotations

import csv
import html
import io
import json
import math
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, unquote, urlparse

import httpx
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import text

from app.ai import complete_chat, embed_texts, list_models_for_tenant, stream_chat_completion
from app.config import settings
from app.deps import RequestContext, get_request_context, require_feature
from app.schemas import NewChat, RenameChat, SendMessage, SetModel
from app.storage import get_storage


router = APIRouter()
WEB_SEARCH_ENDPOINT = "https://duckduckgo.com/html/"
MAX_TABLE_PREVIEW_ROWS = 500
MAX_TABLE_SOURCE_BYTES = 8 * 1024 * 1024
MAX_DOC_EXTRACT_CHARS = 120_000
RAG_CHUNK_SIZE = 1400
RAG_CHUNK_OVERLAP = 220
RAG_MAX_CHUNKS_PER_FILE = 80
RAG_TOP_K = 6
RFP_HISTORY_FILE_KIND = "rfp_history_file"
TAG_RE = re.compile(r"@([A-Za-z0-9][\w.\-]{0,100})")
RECENT_FILE_WORDS = ("file", "files", "document", "documents", "attachment", "attachments", "pdf", "doc", "docx", "sheet", "excel", "csv")
RECENT_UPLOAD_WORDS = ("upload", "uploaded", "attach", "attached", "added")
RECENT_INTENT_WORDS = ("latest", "last", "recent", "new", "just", "today", "now", "right now")


def _sse_event(event: str, payload: Dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(payload, ensure_ascii=True)}\n\n"


def _serialize_datetime(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.isoformat()
    return None


def _serialize_chat(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "title": row.get("title") or "New chat",
        "model": row.get("model") or "",
        "created_at": _serialize_datetime(row.get("created_at")),
        "updated_at": _serialize_datetime(row.get("updated_at")),
    }


def _serialize_message(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row["id"],
        "role": row["role"],
        "content": row["content"],
        "ts": _serialize_datetime(row.get("ts")),
    }


def _strip_html_tags(value: str) -> str:
    no_tags = re.sub(r"<[^>]+>", " ", value)
    return " ".join(html.unescape(no_tags).split())


def _chunk_text(value: str, max_chars: int = 260) -> List[str]:
    text_value = (value or "").strip()
    if not text_value:
        return []

    words = text_value.split()
    chunks: List[str] = []
    current = ""

    for word in words:
        candidate = word if not current else f"{current} {word}"
        if len(candidate) > max_chars and current:
            chunks.append(current)
            current = word
        else:
            current = candidate

    if current:
        chunks.append(current)

    return chunks


def _resolve_search_url(raw_url: str) -> str:
    parsed = urlparse(raw_url)
    if "duckduckgo.com" not in parsed.netloc:
        return raw_url

    query = parse_qs(parsed.query)
    target = query.get("uddg")
    if target and target[0]:
        return unquote(target[0])
    return raw_url


async def _duckduckgo_search(query: str, max_results: int) -> List[Dict[str, str]]:
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; EKChatBot/1.0)",
        "Accept": "text/html,application/xhtml+xml",
    }

    async with httpx.AsyncClient(timeout=settings.EKCHAT_WEBSEARCH_TIMEOUT_SECONDS, follow_redirects=True) as client:
        response = await client.get(WEB_SEARCH_ENDPOINT, params={"q": query}, headers=headers)
        response.raise_for_status()
        html_body = response.text

    link_matches = re.findall(
        r'<a[^>]*class="result__a"[^>]*href="([^"]+)"[^>]*>(.*?)</a>',
        html_body,
        flags=re.IGNORECASE | re.DOTALL,
    )
    snippet_matches = re.findall(
        r'<a[^>]*class="result__snippet"[^>]*>(.*?)</a>|<div[^>]*class="result__snippet"[^>]*>(.*?)</div>',
        html_body,
        flags=re.IGNORECASE | re.DOTALL,
    )

    snippets: List[str] = []
    for a, b in snippet_matches:
        snippets.append(_strip_html_tags(a or b or ""))

    results: List[Dict[str, str]] = []
    for index, (url, title_html) in enumerate(link_matches):
        title = _strip_html_tags(title_html)
        if not title:
            continue

        snippet = snippets[index] if index < len(snippets) else ""
        results.append(
            {
                "title": title,
                "url": _resolve_search_url(url),
                "snippet": snippet,
            }
        )

        if len(results) >= max_results:
            break

    return results


def _fallback_web_answer(query: str, sources: List[Dict[str, str]]) -> str:
    lines = [f"I looked up '{query}'. Here are the top findings:"]
    for idx, source in enumerate(sources[:5], start=1):
        summary = source.get("snippet") or "No snippet was available for this result."
        lines.append(f"{idx}. {source.get('title', 'Untitled')}\n   {summary}")

    return "\n\n".join(lines)


def _decode_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _as_json_dict(value: Any) -> Dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        text_value = value.strip()
        if not text_value:
            return {}
        try:
            parsed = json.loads(text_value)
        except json.JSONDecodeError:
            return {}
        return parsed if isinstance(parsed, dict) else {}
    return {}


def _normalize_embedding(vector: Any, *, target_dim: Optional[int] = None) -> List[float]:
    if not isinstance(vector, list):
        return []
    cleaned: List[float] = []
    for item in vector:
        try:
            value = float(item)
        except (TypeError, ValueError):
            continue
        if math.isfinite(value):
            cleaned.append(value)
    if not cleaned:
        return []
    if target_dim is not None:
        if len(cleaned) > target_dim:
            cleaned = cleaned[:target_dim]
        elif len(cleaned) < target_dim:
            cleaned = cleaned + [0.0] * (target_dim - len(cleaned))
    return cleaned


def _embedding_to_vector_literal(vector: List[float]) -> str:
    return "[" + ",".join(f"{value:.8f}" for value in vector) + "]"


def _cosine_similarity(left: List[float], right: List[float]) -> float:
    if not left or not right:
        return 0.0
    size = min(len(left), len(right))
    if size == 0:
        return 0.0

    dot = 0.0
    left_norm = 0.0
    right_norm = 0.0
    for i in range(size):
        lv = left[i]
        rv = right[i]
        dot += lv * rv
        left_norm += lv * lv
        right_norm += rv * rv

    if left_norm <= 0 or right_norm <= 0:
        return 0.0
    return dot / (math.sqrt(left_norm) * math.sqrt(right_norm))


def _extract_text_for_rag(filename: str, raw: bytes, *, max_chars: int = MAX_DOC_EXTRACT_CHARS) -> str:
    suffix = Path(filename).suffix.lower()

    if suffix == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(raw))
            text_value = "\n".join((page.extract_text() or "") for page in reader.pages)
            return text_value.strip()[:max_chars]
        except Exception:
            return ""

    if suffix == ".docx":
        try:
            from docx import Document

            document = Document(io.BytesIO(raw))
            text_value = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text_value.strip()[:max_chars]
        except Exception:
            return ""

    if suffix == ".json":
        try:
            payload = json.loads(_decode_bytes(raw))
            if isinstance(payload, dict):
                normalized = json.dumps(payload, ensure_ascii=False)
            elif isinstance(payload, list):
                normalized = json.dumps(payload[:300], ensure_ascii=False)
            else:
                normalized = str(payload)
            return normalized.strip()[:max_chars]
        except Exception:
            return _decode_bytes(raw).strip()[:max_chars]

    # Plain text and table-like files still provide useful retrieval context.
    return _decode_bytes(raw).strip()[:max_chars]


def _chunk_document_text(
    text_value: str,
    *,
    chunk_size: int = RAG_CHUNK_SIZE,
    overlap: int = RAG_CHUNK_OVERLAP,
    max_chunks: int = RAG_MAX_CHUNKS_PER_FILE,
) -> List[str]:
    cleaned = re.sub(r"\r\n?", "\n", text_value or "")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    if not cleaned:
        return []

    chunks: List[str] = []
    start = 0
    length = len(cleaned)
    min_break = max(1, chunk_size // 2)

    while start < length and len(chunks) < max_chunks:
        end = min(length, start + chunk_size)
        if end < length:
            window_start = start + min_break
            split_at = max(
                cleaned.rfind("\n\n", window_start, end),
                cleaned.rfind(". ", window_start, end),
                cleaned.rfind(" ", window_start, end),
            )
            if split_at > start:
                end = split_at + 1

        chunk = cleaned[start:end].strip()
        if chunk:
            chunks.append(chunk)

        if end >= length:
            break
        start = max(end - overlap, start + 1)

    return chunks


async def _embedding_udt_name(ctx: RequestContext) -> str:
    result = await ctx.db.execute(
        text(
            """
            SELECT udt_name
            FROM information_schema.columns
            WHERE table_schema = 'ekchat'
              AND table_name = 'vector_chunks'
              AND column_name = 'embedding'
            LIMIT 1
            """
        )
    )
    return str(result.scalar_one_or_none() or "")


async def _insert_vector_chunk(
    ctx: RequestContext,
    *,
    chat_id: str,
    source_id: str,
    chunk_index: int,
    content: str,
    embedding: Optional[List[float]],
    embedding_udt_name: str,
    metadata: Dict[str, Any],
) -> None:
    base_params = {
        "tenant_id": ctx.user.tenant_id,
        "user_id": ctx.user.user_id,
        "chat_id": chat_id,
        "source_id": source_id,
        "chunk_index": chunk_index,
        "content": content,
        "metadata": json.dumps(metadata),
    }

    if embedding and embedding_udt_name == "vector":
        await ctx.db.execute(
            text(
                """
                INSERT INTO ekchat.vector_chunks (
                    tenant_id, user_id, chat_id, source_id, chunk_index, content, embedding, metadata, created_at
                )
                VALUES (
                    :tenant_id,
                    :user_id,
                    :chat_id,
                    :source_id,
                    :chunk_index,
                    :content,
                    CAST(:embedding AS vector),
                    CAST(:metadata AS JSONB),
                    NOW()
                )
                """
            ),
            {
                **base_params,
                "embedding": _embedding_to_vector_literal(embedding),
            },
        )
        return

    if embedding:
        await ctx.db.execute(
            text(
                """
                INSERT INTO ekchat.vector_chunks (
                    tenant_id, user_id, chat_id, source_id, chunk_index, content, embedding, metadata, created_at
                )
                VALUES (
                    :tenant_id,
                    :user_id,
                    :chat_id,
                    :source_id,
                    :chunk_index,
                    :content,
                    CAST(:embedding AS DOUBLE PRECISION[]),
                    CAST(:metadata AS JSONB),
                    NOW()
                )
                """
            ),
            {
                **base_params,
                "embedding": embedding,
            },
        )
        return

    await ctx.db.execute(
        text(
            """
            INSERT INTO ekchat.vector_chunks (
                tenant_id, user_id, chat_id, source_id, chunk_index, content, metadata, created_at
            )
            VALUES (
                :tenant_id,
                :user_id,
                :chat_id,
                :source_id,
                :chunk_index,
                :content,
                CAST(:metadata AS JSONB),
                NOW()
            )
            """
        ),
        base_params,
    )


async def _index_file_for_rag(
    ctx: RequestContext,
    *,
    chat_id: str,
    file_row: Dict[str, Any],
    raw: Optional[bytes] = None,
) -> Dict[str, Any]:
    storage = get_storage()
    file_id = str(file_row["id"])
    file_name = str(file_row.get("original_name") or file_row.get("name") or "document")

    if raw is None:
        raw = await storage.download_bytes(str(file_row["blob_path"]))

    text_value = _extract_text_for_rag(file_name, raw, max_chars=MAX_DOC_EXTRACT_CHARS)
    chunks = _chunk_document_text(text_value)

    await ctx.db.execute(
        text(
            """
            DELETE FROM ekchat.vector_chunks
            WHERE tenant_id = :tenant_id
              AND user_id = :user_id
              AND chat_id = :chat_id
              AND source_id = :source_id
            """
        ),
        {
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "chat_id": chat_id,
            "source_id": file_id,
        },
    )

    if not chunks:
        return {"chunks": 0, "embedded": 0, "warning": "No readable text extracted"}

    embedding_udt = await _embedding_udt_name(ctx)

    vectors: List[List[float]] = []
    embed_warning = ""
    try:
        vectors = await embed_texts(ctx.db, ctx.user.tenant_id, chunks)
    except Exception as exc:
        embed_warning = str(exc)
        vectors = []

    inserted = 0
    embedded = 0
    for idx, chunk in enumerate(chunks):
        vector = vectors[idx] if idx < len(vectors) else None
        if embedding_udt == "vector":
            vector = _normalize_embedding(vector, target_dim=1536)
        else:
            vector = _normalize_embedding(vector)

        await _insert_vector_chunk(
            ctx,
            chat_id=chat_id,
            source_id=file_id,
            chunk_index=idx,
            content=chunk,
            embedding=vector if vector else None,
            embedding_udt_name=embedding_udt,
            metadata={
                "source_id": file_id,
                "source_name": file_name,
                "chunk_index": idx,
            },
        )
        inserted += 1
        if vector:
            embedded += 1

    payload = {"chunks": inserted, "embedded": embedded}
    if embed_warning:
        payload["warning"] = f"Embeddings unavailable; using lexical retrieval fallback ({embed_warning})"
    return payload


async def _list_history_scope_files(
    ctx: RequestContext,
    *,
    source_names: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    names = {str(name).strip() for name in (source_names or []) if str(name).strip()}

    result = await ctx.db.execute(
        text(
            """
            SELECT id, payload, created_at
            FROM ekchat.rfp_sessions
            WHERE tenant_id = :tenant_id
              AND user_id = :user_id
              AND kind = :kind
            ORDER BY created_at DESC
            """
        ),
        {
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "kind": RFP_HISTORY_FILE_KIND,
        },
    )

    files: List[Dict[str, Any]] = []
    for row in result.mappings().all():
        payload = _as_json_dict(row.get("payload"))
        file_name = str(payload.get("name") or "").strip()
        blob_path = str(payload.get("blob_path") or "").strip()
        if not file_name or not blob_path:
            continue
        if names and file_name not in names:
            continue
        files.append(
            {
                "id": str(row["id"]),
                "original_name": file_name,
                "mime_type": payload.get("mime_type"),
                "blob_path": blob_path,
                "file_size": payload.get("file_size"),
                "created_at": row.get("created_at"),
            }
        )

    return files


async def _list_source_files(
    ctx: RequestContext,
    *,
    chat_id: str,
    source_names: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    names = [str(name).strip() for name in (source_names or []) if str(name).strip()]

    if names:
        result = await ctx.db.execute(
            text(
                """
                SELECT id, original_name, mime_type, blob_path, file_size, created_at
                FROM ekchat.chat_files
                WHERE chat_id = :chat_id
                  AND tenant_id = :tenant_id
                  AND user_id = :user_id
                  AND original_name = ANY(CAST(:names AS TEXT[]))
                ORDER BY created_at DESC
                """
            ),
            {
                "chat_id": chat_id,
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
                "names": names,
            },
        )
    else:
        result = await ctx.db.execute(
            text(
                """
                SELECT id, original_name, mime_type, blob_path, file_size, created_at
                FROM ekchat.chat_files
                WHERE chat_id = :chat_id
                  AND tenant_id = :tenant_id
                  AND user_id = :user_id
                ORDER BY created_at DESC
                """
            ),
            {
                "chat_id": chat_id,
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
            },
        )

    files = [dict(row) for row in result.mappings().all()]
    include_history_scope = bool(names) or chat_id.startswith("rfp-history-")
    history_files = (
        await _list_history_scope_files(ctx, source_names=names if names else None)
        if include_history_scope
        else []
    )

    merged_by_id: Dict[str, Dict[str, Any]] = {}
    for row in files + history_files:
        merged_by_id[str(row["id"])] = row
    return list(merged_by_id.values())


async def _ensure_vectors_for_sources(
    ctx: RequestContext,
    *,
    chat_id: str,
    source_rows: List[Dict[str, Any]],
) -> None:
    source_ids = [str(row["id"]) for row in source_rows if row.get("id")]
    if not source_ids:
        return

    count_result = await ctx.db.execute(
        text(
            """
            SELECT source_id, COUNT(*) AS chunk_count
            FROM ekchat.vector_chunks
            WHERE tenant_id = :tenant_id
              AND user_id = :user_id
              AND chat_id = :chat_id
              AND source_id = ANY(CAST(:source_ids AS TEXT[]))
            GROUP BY source_id
            """
        ),
        {
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "chat_id": chat_id,
            "source_ids": source_ids,
        },
    )
    counts = {str(row["source_id"]): int(row["chunk_count"] or 0) for row in count_result.mappings().all()}
    missing_ids = {source_id for source_id in source_ids if counts.get(source_id, 0) == 0}
    if not missing_ids:
        return

    for row in source_rows:
        if str(row.get("id")) not in missing_ids:
            continue
        try:
            await _index_file_for_rag(ctx, chat_id=chat_id, file_row=row)
        except Exception:
            # Retrieval will still attempt lexical fallback from any available rows.
            continue


def _rag_tokenize(value: str) -> List[str]:
    return [token for token in re.split(r"[^A-Za-z0-9]+", (value or "").lower()) if len(token) > 2]


def _lexical_tokens_for_text(value: str) -> List[str]:
    base = _rag_tokenize(value)
    stop = {
        "the", "a", "an", "and", "or", "to", "of", "in", "on", "for", "with",
        "is", "are", "was", "were", "be", "as", "at", "by", "from", "it", "this",
        "that", "these", "those", "about", "what", "who", "when", "where", "why", "how",
        "me", "my", "your", "their", "our",
    }
    return sorted({token for token in base if token and token not in stop})


def _lexical_similarity_score(text_value: str, query: str) -> float:
    tokens = _lexical_tokens_for_text(query)
    if not tokens:
        return 0.0

    haystack = (text_value or "").lower()
    score = sum(1 for token in tokens if token in haystack)
    return min(1.0, score / max(1, len(tokens)))


def _source_slug(value: str) -> str:
    stem = Path(value or "").stem.lower()
    return re.sub(r"[^a-z0-9]+", "", stem)


def _source_rows_sorted_recent(source_rows: List[Dict[str, Any]]) -> List[str]:
    ranked: List[Tuple[float, str]] = []
    for row in source_rows:
        name = str(row.get("original_name") or row.get("name") or "").strip()
        if not name:
            continue

        created = row.get("created_at")
        if isinstance(created, datetime):
            ts = created.timestamp()
        elif isinstance(created, str):
            try:
                ts = datetime.fromisoformat(created.replace("Z", "+00:00")).timestamp()
            except Exception:
                ts = 0.0
        else:
            ts = 0.0

        ranked.append((ts, name))

    ranked.sort(key=lambda item: item[0], reverse=True)

    out: List[str] = []
    seen = set()
    for _, name in ranked:
        if name in seen:
            continue
        seen.add(name)
        out.append(name)
    return out


def _infer_recent_upload_sources(source_rows: List[Dict[str, Any]], text_value: str) -> List[str]:
    content = (text_value or "").lower()
    if not content:
        return []

    has_file_word = any(token in content for token in RECENT_FILE_WORDS)
    has_upload_word = any(token in content for token in RECENT_UPLOAD_WORDS)
    if not (has_file_word and has_upload_word):
        return []

    has_recency = any(token in content for token in RECENT_INTENT_WORDS) or "just uploaded" in content or "recently uploaded" in content
    if not has_recency:
        pattern = r"\b(upload(?:ed|ing)?|attach(?:ed|ing)?)\s+(?:this|that)\s+(?:file|document|attachment)s?\b"
        if not re.search(pattern, content):
            return []

    plural = bool(re.search(r"\b(files|documents|attachments|reports|sheets|pdfs|slides)\b", content))
    limit = 3 if plural else 1
    ordered = _source_rows_sorted_recent(source_rows)
    return ordered[:limit]


def _select_sources_for_chat(
    text_value: str,
    explicit_sources: Optional[List[str]],
    source_rows: List[Dict[str, Any]],
) -> List[str]:
    available = []
    seen = set()
    for row in source_rows:
        name = str(row.get("original_name") or row.get("name") or "").strip()
        if not name or name in seen:
            continue
        seen.add(name)
        available.append(name)

    if not available:
        return []

    if explicit_sources:
        explicit = {str(name).strip() for name in explicit_sources if str(name).strip()}
        return [name for name in available if name in explicit]

    tags = TAG_RE.findall(text_value or "")
    if not tags:
        implied = _infer_recent_upload_sources(source_rows, text_value or "")
        return implied if implied else []

    slug_map = {name: _source_slug(name) for name in available}
    selected: List[str] = []
    for tag in tags:
        tag_slug = _source_slug(tag)
        if not tag_slug:
            continue

        best = None
        for name, slug in slug_map.items():
            if slug == tag_slug:
                best = name
                break
        if not best:
            for name, slug in slug_map.items():
                if slug.startswith(tag_slug):
                    best = name
                    break
        if not best:
            for name, slug in slug_map.items():
                if tag_slug in slug:
                    best = name
                    break

        if best and best not in selected:
            selected.append(best)

    return selected


async def _retrieve_rag_chunks(
    ctx: RequestContext,
    *,
    chat_id: str,
    question: str,
    source_rows: List[Dict[str, Any]],
    top_k: int = RAG_TOP_K,
) -> List[Dict[str, Any]]:
    source_ids = [str(row["id"]) for row in source_rows if row.get("id")]
    if not source_ids:
        return []

    source_name_by_id = {
        str(row["id"]): str(row.get("original_name") or row.get("name") or row.get("id"))
        for row in source_rows
    }

    embedding_udt = await _embedding_udt_name(ctx)

    question_vector: List[float] = []
    try:
        embed_rows = await embed_texts(ctx.db, ctx.user.tenant_id, [question])
        question_vector = _normalize_embedding(embed_rows[0], target_dim=1536 if embedding_udt == "vector" else None)
    except Exception:
        question_vector = []

    search_pool = max(40, top_k * 6)
    candidates: Dict[str, Dict[str, Any]] = {}

    def add_candidate(
        *,
        source_id: str,
        content: str,
        metadata: Any,
        vector_score: float,
        mode: str,
    ) -> None:
        text_value = str(content or "").strip()
        if not text_value:
            return
        key = f"{source_id}|{text_value[:120]}|{len(text_value)}"
        existing = candidates.get(key)
        if existing is None or vector_score > float(existing.get("vector_score") or 0.0):
            candidates[key] = {
                "source_id": source_id,
                "source_name": source_name_by_id.get(source_id, "Document"),
                "content": text_value,
                "vector_score": float(max(0.0, vector_score)),
                "metadata": _as_json_dict(metadata),
                "mode": mode,
            }

    if question_vector and embedding_udt == "vector":
        result = await ctx.db.execute(
            text(
                """
                SELECT source_id, content, metadata,
                       (embedding <-> CAST(:query_embedding AS vector)) AS distance
                FROM ekchat.vector_chunks
                WHERE tenant_id = :tenant_id
                  AND user_id = :user_id
                  AND chat_id = :chat_id
                  AND source_id = ANY(CAST(:source_ids AS TEXT[]))
                  AND embedding IS NOT NULL
                ORDER BY embedding <-> CAST(:query_embedding AS vector)
                LIMIT :limit
                """
            ),
            {
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
                "chat_id": chat_id,
                "source_ids": source_ids,
                "query_embedding": _embedding_to_vector_literal(question_vector),
                "limit": search_pool,
            },
        )
        for row in result.mappings().all():
            distance = float(row.get("distance") or 0.0)
            add_candidate(
                source_id=str(row.get("source_id") or ""),
                content=str(row.get("content") or ""),
                metadata=row.get("metadata"),
                vector_score=1.0 / (1.0 + max(0.0, distance)),
                mode="vector-distance",
            )
    elif question_vector:
        result = await ctx.db.execute(
            text(
                """
                SELECT source_id, content, metadata, embedding
                FROM ekchat.vector_chunks
                WHERE tenant_id = :tenant_id
                  AND user_id = :user_id
                  AND chat_id = :chat_id
                  AND source_id = ANY(CAST(:source_ids AS TEXT[]))
                  AND embedding IS NOT NULL
                LIMIT :limit
                """
            ),
            {
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
                "chat_id": chat_id,
                "source_ids": source_ids,
                "limit": max(200, search_pool * 5),
            },
        )

        for row in result.mappings().all():
            vector = _normalize_embedding(row.get("embedding"))
            score = _cosine_similarity(question_vector, vector)
            if score <= 0:
                continue
            add_candidate(
                source_id=str(row.get("source_id") or ""),
                content=str(row.get("content") or ""),
                metadata=row.get("metadata"),
                vector_score=score,
                mode="vector-cosine",
            )

    # Ekchat-style lexical fallback and blend for better recall on sparse embeddings.
    if len(candidates) < top_k:
        result = await ctx.db.execute(
            text(
                """
                SELECT source_id, content, metadata
                FROM ekchat.vector_chunks
                WHERE tenant_id = :tenant_id
                  AND user_id = :user_id
                  AND chat_id = :chat_id
                  AND source_id = ANY(CAST(:source_ids AS TEXT[]))
                LIMIT 900
                """
            ),
            {
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
                "chat_id": chat_id,
                "source_ids": source_ids,
            },
        )

        for row in result.mappings().all():
            content = str(row.get("content") or "").strip()
            lexical = _lexical_similarity_score(content, question)
            if lexical <= 0:
                continue
            add_candidate(
                source_id=str(row.get("source_id") or ""),
                content=content,
                metadata=row.get("metadata"),
                vector_score=lexical,
                mode="lexical",
            )

    ranked: List[Dict[str, Any]] = []
    for item in candidates.values():
        lexical_score = _lexical_similarity_score(item.get("content") or "", question)
        combined = max(float(item.get("vector_score") or 0.0), lexical_score)
        if combined <= 0:
            continue
        ranked.append(
            {
                "source_id": item["source_id"],
                "source_name": item["source_name"],
                "content": item["content"],
                "score": combined,
                "metadata": item.get("metadata") or {},
            }
        )

    ranked.sort(key=lambda entry: float(entry.get("score") or 0.0), reverse=True)
    return ranked[: max(1, top_k)]


def _build_rag_context(chunks: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for idx, chunk in enumerate(chunks, start=1):
        source_name = str(chunk.get("source_name") or "Document")
        content = str(chunk.get("content") or "").strip()
        if not content:
            continue
        lines.append(f"[{idx}] Source: {source_name}\n{content}")
    return "\n\n".join(lines).strip()


def _coerce_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text_value = str(value).strip().replace(",", "")
    if not text_value:
        return None

    try:
        return float(text_value)
    except ValueError:
        return None


def _build_columns_metadata(headers: List[str], rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    columns: List[Dict[str, str]] = []
    sample_rows = rows[:80]

    for header in headers:
        filled = 0
        numeric = 0
        for row in sample_rows:
            value = row.get(header)
            if value is None or str(value).strip() == "":
                continue
            filled += 1
            if _coerce_number(value) is not None:
                numeric += 1

        is_numeric = filled > 0 and numeric / max(1, filled) >= 0.75
        columns.append({"name": header, "kind": "number" if is_numeric else "other"})

    return columns


def _normalize_headers(raw_headers: List[Any]) -> List[str]:
    headers: List[str] = []
    for idx, value in enumerate(raw_headers):
        candidate = str(value or "").strip()
        headers.append(candidate or f"column_{idx + 1}")
    return headers


def _parse_csv_rows(raw: bytes, *, delimiter: str, limit: int) -> Tuple[List[str], List[Dict[str, Any]], int]:
    text_value = _decode_bytes(raw)
    reader = csv.reader(io.StringIO(text_value), delimiter=delimiter)

    try:
        first_row = next(reader)
    except StopIteration:
        return [], [], 0

    headers = _normalize_headers(list(first_row))
    rows: List[Dict[str, Any]] = []
    total_rows = 0

    for raw_row in reader:
        total_rows += 1
        if len(rows) >= limit:
            continue

        padded = list(raw_row) + [""] * (len(headers) - len(raw_row))
        row_payload = {headers[i]: padded[i] for i in range(len(headers))}
        rows.append(row_payload)

    return headers, rows, total_rows


def _parse_json_rows(raw: bytes, *, limit: int) -> Tuple[List[str], List[Dict[str, Any]], int]:
    text_value = _decode_bytes(raw)

    try:
        payload = json.loads(text_value)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid JSON content: {exc.msg}") from exc

    rows_source: List[Dict[str, Any]] = []

    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                rows_source.append(item)
            else:
                rows_source.append({"value": item})
    elif isinstance(payload, dict):
        list_value = None
        for value in payload.values():
            if isinstance(value, list):
                list_value = value
                break

        if list_value is not None:
            for item in list_value:
                if isinstance(item, dict):
                    rows_source.append(item)
                else:
                    rows_source.append({"value": item})
        else:
            rows_source = [payload]
    else:
        rows_source = [{"value": payload}]

    if not rows_source:
        return [], [], 0

    headers_set = set()
    for item in rows_source:
        headers_set.update(item.keys())

    headers = sorted(headers_set)
    rows: List[Dict[str, Any]] = []

    for item in rows_source[:limit]:
        rows.append({header: item.get(header) for header in headers})

    return headers, rows, len(rows_source)


def _list_xlsx_sheets(raw: bytes) -> List[str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XLSX parsing requires openpyxl: {exc}") from exc

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _parse_xlsx_rows(raw: bytes, *, sheet_name: Optional[str], limit: int) -> Tuple[List[str], List[Dict[str, Any]], int, str]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XLSX parsing requires openpyxl: {exc}") from exc

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

    try:
        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{selected_sheet}' not found")

        worksheet = workbook[selected_sheet]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            first_row = next(iterator)
        except StopIteration:
            return [], [], 0, selected_sheet

        headers = _normalize_headers(list(first_row))
        rows: List[Dict[str, Any]] = []
        total_rows = 0

        for raw_row in iterator:
            total_rows += 1
            if len(rows) >= limit:
                continue

            values = list(raw_row or ())
            values += [""] * (len(headers) - len(values))
            rows.append({headers[i]: values[i] for i in range(len(headers))})

        return headers, rows, total_rows, selected_sheet
    finally:
        workbook.close()


def _is_tabular_file(file_name: str, mime_type: Optional[str]) -> bool:
    suffix = Path(file_name).suffix.lower()
    if suffix in {".csv", ".tsv", ".json", ".xlsx", ".xlsm"}:
        return True

    mime = (mime_type or "").lower()
    return any(
        token in mime
        for token in (
            "text/csv",
            "application/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/json",
        )
    )


def _parse_table_reference(table_ref: str) -> Tuple[str, Optional[str]]:
    if ":" not in table_ref:
        return table_ref, None

    file_id, selector = table_ref.split(":", 1)
    return file_id, selector or None


def _load_table_from_file(
    *,
    file_name: str,
    mime_type: Optional[str],
    raw: bytes,
    selector: Optional[str],
    limit: int,
) -> Dict[str, Any]:
    suffix = Path(file_name).suffix.lower()

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        headers, rows, total = _parse_csv_rows(raw, delimiter=delimiter, limit=limit)
        return {"headers": headers, "rows": rows, "total_rows": total, "sheet": None}

    if suffix == ".json":
        headers, rows, total = _parse_json_rows(raw, limit=limit)
        return {"headers": headers, "rows": rows, "total_rows": total, "sheet": None}

    if suffix in {".xlsx", ".xlsm"}:
        headers, rows, total, sheet = _parse_xlsx_rows(raw, sheet_name=selector, limit=limit)
        return {"headers": headers, "rows": rows, "total_rows": total, "sheet": sheet}

    if mime_type and "json" in mime_type.lower():
        headers, rows, total = _parse_json_rows(raw, limit=limit)
        return {"headers": headers, "rows": rows, "total_rows": total, "sheet": None}

    raise HTTPException(status_code=400, detail="File format is not supported for table preview")


def _safe_artifact_name(value: Optional[str]) -> str:
    candidate = Path(value or "").name.strip()
    candidate = re.sub(r"[^A-Za-z0-9._-]+", "-", candidate)
    candidate = candidate.strip(".-")
    if not candidate:
        candidate = f"plot-{uuid.uuid4().hex[:10]}.png"
    if not candidate.lower().endswith(".png"):
        candidate = f"{candidate}.png"
    return candidate


def _render_plot_png(
    rows: List[Dict[str, Any]],
    *,
    x_column: Optional[str],
    y_column: Optional[str],
    kind: Optional[str],
    title: Optional[str],
) -> Tuple[bytes, str, str, Optional[str]]:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Plotting requires matplotlib: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=400, detail="No rows available for plotting")

    columns = list(rows[0].keys())
    if not columns:
        raise HTTPException(status_code=400, detail="Table has no columns")

    numeric_columns = [
        column
        for column in columns
        if any(_coerce_number(row.get(column)) is not None for row in rows)
    ]

    chosen_kind = (kind or "line").strip().lower()
    if chosen_kind not in {"line", "bar", "scatter", "hist"}:
        chosen_kind = "line"

    fig, ax = plt.subplots(figsize=(9.5, 5.5))

    chosen_x = x_column
    chosen_y = y_column

    if chosen_kind == "hist":
        target_col = chosen_y or chosen_x or (numeric_columns[0] if numeric_columns else None)
        if not target_col:
            raise HTTPException(status_code=400, detail="Histogram requires at least one numeric column")

        values = [_coerce_number(row.get(target_col)) for row in rows]
        numeric_values = [value for value in values if value is not None]
        if not numeric_values:
            raise HTTPException(status_code=400, detail=f"Column '{target_col}' has no numeric data")

        ax.hist(numeric_values, bins=min(20, max(5, len(numeric_values) // 4)))
        ax.set_xlabel(target_col)
        ax.set_ylabel("Count")
        chosen_x = target_col
        chosen_y = None
    else:
        selected_x = chosen_x or columns[0]
        selected_y = chosen_y

        if not selected_y:
            selected_y = next((col for col in numeric_columns if col != selected_x), None)
            if not selected_y and selected_x in numeric_columns:
                selected_y = selected_x

        if not selected_y:
            raise HTTPException(status_code=400, detail="Plot requires a numeric y-axis column")

        x_values: List[Any] = []
        y_values: List[float] = []

        for row in rows:
            y_num = _coerce_number(row.get(selected_y))
            if y_num is None:
                continue
            x_values.append(row.get(selected_x))
            y_values.append(y_num)

        if not y_values:
            raise HTTPException(status_code=400, detail=f"Column '{selected_y}' has no numeric data")

        if chosen_kind == "bar":
            ax.bar([str(value) for value in x_values], y_values)
        elif chosen_kind == "scatter":
            x_numeric = [_coerce_number(value) for value in x_values]
            if all(value is not None for value in x_numeric):
                ax.scatter(x_numeric, y_values)
            else:
                ax.scatter(range(len(y_values)), y_values)
                ax.set_xticks(range(len(y_values)))
                ax.set_xticklabels([str(value) for value in x_values], rotation=30, ha="right")
        else:
            ax.plot([str(value) for value in x_values], y_values)

        ax.set_xlabel(selected_x)
        ax.set_ylabel(selected_y)
        chosen_x = selected_x
        chosen_y = selected_y

    ax.set_title((title or f"{chosen_kind.title()} chart").strip())
    fig.tight_layout()

    output = io.BytesIO()
    fig.savefig(output, format="png", dpi=140)
    plt.close(fig)
    return output.getvalue(), chosen_kind, chosen_x or "", chosen_y


async def _load_chat(ctx: RequestContext, chat_id: str) -> Dict[str, Any]:
    result = await ctx.db.execute(
        text(
            """
            SELECT id, tenant_id, user_id, title, model, created_at, updated_at
            FROM ekchat.chats
            WHERE id = :chat_id AND tenant_id = :tenant_id AND user_id = :user_id
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Chat not found")
    return dict(row)


async def _append_message(ctx: RequestContext, chat_id: str, role: str, content: str) -> str:
    message_id = str(uuid.uuid4())
    await ctx.db.execute(
        text(
            """
            INSERT INTO ekchat.messages (id, chat_id, tenant_id, user_id, role, content, ts)
            VALUES (:id, :chat_id, :tenant_id, :user_id, :role, :content, NOW())
            """
        ),
        {
            "id": message_id,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "role": role,
            "content": content,
        },
    )
    await ctx.db.execute(
        text(
            """
            UPDATE ekchat.chats
            SET updated_at = NOW()
            WHERE id = :chat_id AND tenant_id = :tenant_id AND user_id = :user_id
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    return message_id


async def _load_message_history(
    ctx: RequestContext,
    chat_id: str,
    limit: int = settings.EKCHAT_MAX_CONTEXT_MESSAGES,
) -> List[Dict[str, str]]:
    result = await ctx.db.execute(
        text(
            """
            SELECT role, content
            FROM (
                SELECT role, content, ts
                FROM ekchat.messages
                WHERE chat_id = :chat_id
                  AND tenant_id = :tenant_id
                  AND user_id = :user_id
                ORDER BY ts DESC
                LIMIT :limit_count
            ) AS recent
            ORDER BY ts ASC
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "limit_count": limit,
        },
    )
    rows = result.mappings().all()
    return [{"role": row["role"], "content": row["content"]} for row in rows]


async def _load_chat_file_row(ctx: RequestContext, chat_id: str, file_id: str) -> Dict[str, Any]:
    result = await ctx.db.execute(
        text(
            """
            SELECT id, original_name, mime_type, blob_path, file_size, created_at
            FROM ekchat.chat_files
            WHERE id = :file_id
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            LIMIT 1
            """
        ),
        {
            "file_id": file_id,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="File not found")
    return dict(row)


@router.get("/health")
async def health() -> Dict[str, Any]:
    return {"ok": True, "service": "ekchat-api", "time": datetime.utcnow().isoformat()}


@router.get("/whoami")
async def whoami(ctx: RequestContext = Depends(get_request_context)):
    return {
        "user_id": ctx.user.user_id,
        "tenant_id": ctx.user.tenant_id,
        "email": ctx.user.email,
        "role": ctx.user.role,
    }


@router.get("/models")
async def models(ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False))):
    return {"models": await list_models_for_tenant(ctx.db, ctx.user.tenant_id)}


@router.post("/models/ensure")
async def models_ensure(
    body: SetModel,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    models = await list_models_for_tenant(ctx.db, ctx.user.tenant_id)
    return {
        "status": "ok",
        "requested": body.model,
        "available": body.model in models,
        "models": models,
    }


@router.get("/chats")
async def list_chats(ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False))):
    result = await ctx.db.execute(
        text(
            """
            SELECT id, title, model, created_at, updated_at
            FROM ekchat.chats
            WHERE tenant_id = :tenant_id AND user_id = :user_id
            ORDER BY updated_at DESC
            """
        ),
        {"tenant_id": ctx.user.tenant_id, "user_id": ctx.user.user_id},
    )
    rows = result.mappings().all()
    return {"chats": [_serialize_chat(dict(row)) for row in rows]}


@router.post("/chats")
async def new_chat(
    body: NewChat,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    chat_id = str(uuid.uuid4())
    await ctx.db.execute(
        text(
            """
            INSERT INTO ekchat.chats (id, tenant_id, user_id, title, model, created_at, updated_at)
            VALUES (:id, :tenant_id, :user_id, :title, :model, NOW(), NOW())
            """
        ),
        {
            "id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "title": body.title or "New chat",
            "model": body.model,
        },
    )
    chat = await _load_chat(ctx, chat_id)
    return {"chat": _serialize_chat(chat)}


@router.get("/chats/{chat_id}")
async def get_chat_meta(
    chat_id: str,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    chat = await _load_chat(ctx, chat_id)
    return {"chat": _serialize_chat(chat)}


@router.get("/chats/{chat_id}/messages")
async def get_chat_messages(
    chat_id: str,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    result = await ctx.db.execute(
        text(
            """
            SELECT id, role, content, ts
            FROM ekchat.messages
            WHERE chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY ts ASC
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    rows = result.mappings().all()
    return {"messages": [_serialize_message(dict(row)) for row in rows]}


@router.post("/chats/{chat_id}/model")
async def update_chat_model(
    chat_id: str,
    body: SetModel,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    await ctx.db.execute(
        text(
            """
            UPDATE ekchat.chats
            SET model = :model, updated_at = NOW()
            WHERE id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "model": body.model,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    return {"ok": True, "chat": _serialize_chat(await _load_chat(ctx, chat_id))}


@router.post("/chats/{chat_id}/title")
async def rename_chat(
    chat_id: str,
    body: RenameChat,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    await ctx.db.execute(
        text(
            """
            UPDATE ekchat.chats
            SET title = :title, updated_at = NOW()
            WHERE id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "title": body.title.strip(),
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    return {"ok": True, "chat": _serialize_chat(await _load_chat(ctx, chat_id))}


@router.post("/chats/{chat_id}/title/auto")
async def auto_title_chat(
    chat_id: str,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    result = await ctx.db.execute(
        text(
            """
            SELECT content
            FROM ekchat.messages
            WHERE chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
              AND role = 'user'
            ORDER BY ts ASC
            LIMIT 1
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    first_user = result.scalar_one_or_none() or "New chat"
    title = " ".join(first_user.strip().split())[:80] or "New chat"

    await ctx.db.execute(
        text(
            """
            UPDATE ekchat.chats
            SET title = :title, updated_at = NOW()
            WHERE id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "title": title,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    return {"ok": True, "chat": _serialize_chat(await _load_chat(ctx, chat_id))}


@router.post("/chats/{chat_id}/message")
async def send_message(
    chat_id: str,
    body: SendMessage,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    chat = await _load_chat(ctx, chat_id)
    model_name = body.model or chat.get("model")

    if not model_name:
        models = await list_models_for_tenant(ctx.db, ctx.user.tenant_id)
        if not models:
            raise HTTPException(status_code=400, detail="No model configured for tenant")
        model_name = models[0]

    await _append_message(ctx, chat_id, "user", body.content)

    history = await _load_message_history(ctx, chat_id)

    all_source_rows = await _list_source_files(ctx, chat_id=chat_id, source_names=None)
    selected_names = _select_sources_for_chat(
        body.content,
        body.sources,
        all_source_rows,
    )

    is_doc_chat = chat_id.startswith("rfp-history-")
    if not selected_names and is_doc_chat:
        selected_names = [
            str(row.get("original_name") or row.get("name") or "").strip()
            for row in all_source_rows
            if str(row.get("original_name") or row.get("name") or "").strip()
        ]
        if not selected_names:
            selected_names = ["__none__"]

    if selected_names and "__none__" in selected_names:
        source_rows: List[Dict[str, Any]] = []
    elif selected_names:
        selected_set = set(selected_names)
        source_rows = [
            row
            for row in all_source_rows
            if str(row.get("original_name") or row.get("name") or "").strip() in selected_set
        ]
    else:
        source_rows = []

    rag_context = ""
    if source_rows:
        await _ensure_vectors_for_sources(ctx, chat_id=chat_id, source_rows=source_rows)
        rag_chunks = await _retrieve_rag_chunks(
            ctx,
            chat_id=chat_id,
            question=body.content,
            source_rows=source_rows,
            top_k=RAG_TOP_K,
        )
        rag_context = _build_rag_context(rag_chunks)

    model_history = list(history)
    if rag_context and model_history:
        latest = model_history[-1]
        if latest.get("role") == "user":
            question = str(latest.get("content") or "").strip()
            if is_doc_chat:
                system_content = (
                    "You are a proposal assistant. Use only the provided document context for factual claims. "
                    "If the answer is missing from context, say so briefly and ask a clarifying question. "
                    "When values differ across proposals, explicitly state they vary by solicitation and list available values. "
                    "Never answer 'not specified' unless the value is absent from all provided sources. "
                    "Do not include citations or a Sources section unless the user explicitly asks."
                )
            else:
                system_content = (
                    "You are a retrieval-augmented assistant. Prioritize supplied document context over assumptions. "
                    "If context is insufficient, say so clearly."
                )

            model_history[-1] = {
                "role": "user",
                "content": (
                    "Use the provided document context to answer the user question. "
                    "If the answer is not in the context, say you cannot find it in the uploaded documents.\n\n"
                    f"Document context:\n{rag_context}\n\n"
                    f"User question:\n{question}"
                ),
            }
            model_history.insert(0, {"role": "system", "content": system_content})

    async def event_stream():
        assistant_chunks: List[str] = []
        try:
            yield _sse_event("meta", {"chat_id": chat_id, "model": model_name})
            async for token in stream_chat_completion(
                ctx.db,
                ctx.user.tenant_id,
                model_history,
                model_name=model_name,
            ):
                assistant_chunks.append(token)
                yield _sse_event("delta", {"content": token})

            final_text = "".join(assistant_chunks).strip()
            if not final_text:
                final_text = "The model did not return any text."

            await _append_message(ctx, chat_id, "assistant", final_text)
            yield _sse_event("done", {"content": final_text})
        except Exception as exc:
            yield _sse_event("error", {"message": str(exc)})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@router.post("/chats/{chat_id}/websearch")
async def websearch(
    chat_id: str,
    body: SendMessage,
    ctx: RequestContext = Depends(require_feature("ekchat_rag_enabled", default=False)),
):
    chat = await _load_chat(ctx, chat_id)

    query = (body.content or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="Query content is required")

    model_name = body.model or chat.get("model") or settings.EKCHAT_LIGHT_TASK_MODEL
    await _append_message(ctx, chat_id, "user", query)

    try:
        sources = await _duckduckgo_search(query, max_results=settings.EKCHAT_WEBSEARCH_MAX_RESULTS)
    except Exception as exc:
        sources = []
        fallback_error = f"Web search failed: {exc}"
    else:
        fallback_error = ""

    if not sources:
        if fallback_error:
            final_text = f"I couldn't fetch web results right now. {fallback_error}"
        else:
            final_text = "I couldn't find web results for that query right now."
    else:
        source_lines = []
        for idx, source in enumerate(sources, start=1):
            snippet = source.get("snippet") or "No snippet provided."
            source_lines.append(
                f"[{idx}] {source.get('title', 'Untitled')}\n"
                f"URL: {source.get('url', '')}\n"
                f"Snippet: {snippet}"
            )

        system_prompt = (
            "You are a concise analyst. Use only the provided web sources. "
            "Do not invent links. Mention uncertainty clearly when sources are weak."
        )
        user_prompt = (
            f"Question: {query}\n\n"
            f"Sources:\n{chr(10).join(source_lines)}\n\n"
            "Write a useful answer in plain language and end with a short bullet list titled Sources."
        )

        try:
            final_text = await complete_chat(
                ctx.db,
                ctx.user.tenant_id,
                [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                model_name=model_name,
                max_tokens=max(500, settings.EKCHAT_CHAT_MAX_TOKENS),
                temperature=0.2,
            )
            final_text = (final_text or "").strip()
        except Exception:
            final_text = ""

        if not final_text:
            final_text = _fallback_web_answer(query, sources)

        if "Sources:" not in final_text:
            refs = "\n".join(
                f"- [{idx}] {source.get('title', 'Untitled')} ({source.get('url', '')})"
                for idx, source in enumerate(sources, start=1)
            )
            final_text = f"{final_text.strip()}\n\nSources:\n{refs}".strip()

    await _append_message(ctx, chat_id, "assistant", final_text)

    async def event_stream():
        yield _sse_event("meta", {"chat_id": chat_id, "model": model_name, "mode": "websearch"})
        for chunk in _chunk_text(final_text):
            yield _sse_event("delta", {"content": chunk})
        yield _sse_event("done", {"content": final_text})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@router.delete("/chats/{chat_id}")
async def delete_chat(
    chat_id: str,
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    await ctx.db.execute(
        text(
            """
            DELETE FROM ekchat.chats
            WHERE id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    return {"ok": True}


@router.post("/files/upload")
async def upload_file(
    chat_id: str = Form(...),
    file: Optional[UploadFile] = File(default=None),
    files: List[UploadFile] = File(default=[]),
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)

    upload_items = [item for item in ([file] + list(files)) if item is not None]
    if not upload_items:
        raise HTTPException(status_code=400, detail="No file provided")

    storage = get_storage()
    uploaded_files: List[Dict[str, Any]] = []

    for upload in upload_items:
        raw = await upload.read()
        if not raw:
            continue

        safe_name = Path(upload.filename or "upload.bin").name
        file_id = str(uuid.uuid4())
        blob_path = (
            f"tenant/{ctx.user.tenant_id}/user/{ctx.user.user_id}/chat/{chat_id}/files/"
            f"{file_id}-{safe_name}"
        )

        await storage.upload_bytes(blob_path, raw, content_type=upload.content_type)

        await ctx.db.execute(
            text(
                """
                INSERT INTO ekchat.chat_files (
                    id,
                    chat_id,
                    tenant_id,
                    user_id,
                    original_name,
                    mime_type,
                    blob_path,
                    file_size,
                    created_at
                )
                VALUES (
                    :id,
                    :chat_id,
                    :tenant_id,
                    :user_id,
                    :original_name,
                    :mime_type,
                    :blob_path,
                    :file_size,
                    NOW()
                )
                """
            ),
            {
                "id": file_id,
                "chat_id": chat_id,
                "tenant_id": ctx.user.tenant_id,
                "user_id": ctx.user.user_id,
                "original_name": safe_name,
                "mime_type": upload.content_type,
                "blob_path": blob_path,
                "file_size": len(raw),
            },
        )

        index_status = await _index_file_for_rag(
            ctx,
            chat_id=chat_id,
            file_row={
                "id": file_id,
                "original_name": safe_name,
                "mime_type": upload.content_type,
                "blob_path": blob_path,
            },
            raw=raw,
        )

        uploaded_files.append(
            {
                "id": file_id,
                "name": safe_name,
                "mime_type": upload.content_type,
                "size": len(raw),
                "blob_path": blob_path,
                "indexing": index_status,
            }
        )

    if not uploaded_files:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    return {
        "ok": True,
        "message": "Files indexed.",
        "file": uploaded_files[0],
        "files": uploaded_files,
    }


@router.post("/files/delete")
async def delete_file(
    body: Dict[str, Any] = Body(...),
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    chat_id = str(body.get("chat_id") or "").strip()
    name = str(body.get("name") or "").strip()
    file_id = str(body.get("file_id") or "").strip()

    if not chat_id:
        raise HTTPException(status_code=400, detail="chat_id is required")
    if not name and not file_id:
        raise HTTPException(status_code=400, detail="Either name or file_id is required")

    await _load_chat(ctx, chat_id)

    if file_id:
        query = text(
            """
            SELECT id, blob_path
            FROM ekchat.chat_files
            WHERE id = :id
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            LIMIT 1
            """
        )
        params = {
            "id": file_id,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        }
    else:
        query = text(
            """
            SELECT id, blob_path
            FROM ekchat.chat_files
            WHERE original_name = :name
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY created_at DESC
            LIMIT 1
            """
        )
        params = {
            "name": name,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        }

    result = await ctx.db.execute(query, params)
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="File not found")

    storage = get_storage()
    try:
        await storage.delete_path(str(row["blob_path"]))
    except Exception:
        pass

    await ctx.db.execute(
        text(
            """
            DELETE FROM ekchat.chat_files
            WHERE id = :id
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "id": row["id"],
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )
    await ctx.db.execute(
        text(
            """
            DELETE FROM ekchat.vector_chunks
            WHERE source_id = :source_id
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            """
        ),
        {
            "source_id": row["id"],
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    return {"ok": True}


@router.get("/files/list")
async def list_files(
    chat_id: str = Query(...),
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)

    result = await ctx.db.execute(
        text(
            """
            SELECT id, original_name, mime_type, file_size, created_at
            FROM ekchat.chat_files
            WHERE chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY created_at DESC
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    files = []
    for row in result.mappings().all():
        files.append(
            {
                "id": row["id"],
                "name": row["original_name"],
                "mime_type": row["mime_type"],
                "size": row["file_size"],
                "created_at": _serialize_datetime(row.get("created_at")),
            }
        )

    return {"files": files}


@router.get("/files/download")
async def download_file(
    chat_id: str = Query(...),
    name: Optional[str] = Query(default=None),
    file_id: Optional[str] = Query(default=None),
    ctx: RequestContext = Depends(require_feature("ekchat_chat_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)

    if not name and not file_id:
        raise HTTPException(status_code=400, detail="Either 'name' or 'file_id' is required.")

    if file_id:
        query = text(
            """
            SELECT original_name, mime_type, blob_path
            FROM ekchat.chat_files
            WHERE id = :file_id
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            LIMIT 1
            """
        )
        params = {
            "file_id": file_id,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        }
    else:
        query = text(
            """
            SELECT original_name, mime_type, blob_path
            FROM ekchat.chat_files
            WHERE original_name = :name
              AND chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY created_at DESC
            LIMIT 1
            """
        )
        params = {
            "name": name,
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        }

    result = await ctx.db.execute(query, params)
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="File not found")

    storage = get_storage()
    content = await storage.download_bytes(row["blob_path"])
    filename = row["original_name"]

    return Response(
        content=content,
        media_type=row.get("mime_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/data/tables")
async def data_tables(
    chat_id: str = Query(...),
    ctx: RequestContext = Depends(require_feature("ekchat_rag_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)

    result = await ctx.db.execute(
        text(
            """
            SELECT id, original_name, mime_type, blob_path, file_size, created_at
            FROM ekchat.chat_files
            WHERE chat_id = :chat_id
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY created_at DESC
            """
        ),
        {
            "chat_id": chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    storage = get_storage()
    tables: List[Dict[str, Any]] = []

    for row in result.mappings().all():
        file_id = str(row["id"])
        file_name = row["original_name"]
        mime_type = row.get("mime_type")
        file_size = row.get("file_size") or 0

        if not _is_tabular_file(file_name, mime_type):
            continue

        if file_size and file_size > MAX_TABLE_SOURCE_BYTES:
            tables.append(
                {
                    "id": file_id,
                    "name": file_name,
                    "source_file_id": file_id,
                    "columns": [],
                    "n_rows": 0,
                    "parse_error": f"File is too large for preview (>{MAX_TABLE_SOURCE_BYTES} bytes).",
                }
            )
            continue

        try:
            raw = await storage.download_bytes(row["blob_path"])
        except Exception as exc:
            tables.append(
                {
                    "id": file_id,
                    "name": file_name,
                    "source_file_id": file_id,
                    "columns": [],
                    "n_rows": 0,
                    "parse_error": f"Failed to load file from storage: {exc}",
                }
            )
            continue

        suffix = Path(file_name).suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm"}:
                sheet_names = _list_xlsx_sheets(raw)
                for sheet_name in sheet_names:
                    parsed = _load_table_from_file(
                        file_name=file_name,
                        mime_type=mime_type,
                        raw=raw,
                        selector=sheet_name,
                        limit=80,
                    )
                    headers = parsed["headers"]
                    rows = parsed["rows"]
                    tables.append(
                        {
                            "id": f"{file_id}:{sheet_name}",
                            "name": f"{file_name} ({sheet_name})",
                            "source_file_id": file_id,
                            "sheet": sheet_name,
                            "columns": _build_columns_metadata(headers, rows),
                            "n_rows": parsed["total_rows"],
                        }
                    )
            else:
                parsed = _load_table_from_file(
                    file_name=file_name,
                    mime_type=mime_type,
                    raw=raw,
                    selector=None,
                    limit=80,
                )
                headers = parsed["headers"]
                rows = parsed["rows"]
                tables.append(
                    {
                        "id": file_id,
                        "name": file_name,
                        "source_file_id": file_id,
                        "columns": _build_columns_metadata(headers, rows),
                        "n_rows": parsed["total_rows"],
                    }
                )
        except HTTPException as exc:
            tables.append(
                {
                    "id": file_id,
                    "name": file_name,
                    "source_file_id": file_id,
                    "columns": [],
                    "n_rows": 0,
                    "parse_error": exc.detail,
                }
            )

    return {"tables": tables}


@router.get("/data/preview")
async def data_preview(
    chat_id: str = Query(...),
    table: str = Query(...),
    limit: int = Query(default=20, ge=1, le=200),
    ctx: RequestContext = Depends(require_feature("ekchat_rag_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)
    file_id, selector = _parse_table_reference(table)

    file_row = await _load_chat_file_row(ctx, chat_id, file_id)
    storage = get_storage()
    raw = await storage.download_bytes(file_row["blob_path"])

    parsed = _load_table_from_file(
        file_name=file_row["original_name"],
        mime_type=file_row.get("mime_type"),
        raw=raw,
        selector=selector,
        limit=min(MAX_TABLE_PREVIEW_ROWS, limit),
    )

    headers = parsed["headers"]
    rows = parsed["rows"][:limit]

    return {
        "table_id": table,
        "columns": _build_columns_metadata(headers, rows),
        "rows": rows,
        "n_rows": parsed["total_rows"],
        "sheet": parsed.get("sheet"),
    }


@router.post("/data/plot")
async def data_plot(
    chat_id: Optional[str] = Query(default=None),
    table: Optional[str] = Query(default=None),
    x: Optional[str] = Query(default=None),
    y: Optional[str] = Query(default=None),
    kind: Optional[str] = Query(default=None),
    name: Optional[str] = Query(default=None),
    payload: Optional[Dict[str, Any]] = Body(default=None),
    ctx: RequestContext = Depends(require_feature("ekchat_rag_enabled", default=False)),
):
    body = payload or {}

    resolved_chat_id = chat_id or body.get("chat_id")
    table_ref = table or body.get("table") or body.get("table_id")
    x_value = x if x is not None else body.get("x")
    y_value = y if y is not None else body.get("y")
    kind_value = kind if kind is not None else body.get("kind")
    name_value = name if name is not None else body.get("name")

    if not resolved_chat_id or not table_ref:
        raise HTTPException(status_code=400, detail="chat_id and table are required")

    await _load_chat(ctx, resolved_chat_id)

    file_id, selector = _parse_table_reference(str(table_ref))
    file_row = await _load_chat_file_row(ctx, resolved_chat_id, file_id)

    storage = get_storage()
    raw = await storage.download_bytes(file_row["blob_path"])
    parsed = _load_table_from_file(
        file_name=file_row["original_name"],
        mime_type=file_row.get("mime_type"),
        raw=raw,
        selector=selector,
        limit=1000,
    )

    plot_bytes, chosen_kind, chosen_x, chosen_y = _render_plot_png(
        parsed["rows"],
        x_column=x_value,
        y_column=y_value,
        kind=kind_value,
        title=body.get("title"),
    )

    artifact_id = str(uuid.uuid4())
    artifact_name = _safe_artifact_name(name_value)
    blob_path = (
        f"tenant/{ctx.user.tenant_id}/user/{ctx.user.user_id}/chat/{resolved_chat_id}/plots/"
        f"{artifact_id}-{artifact_name}"
    )

    await storage.upload_bytes(blob_path, plot_bytes, content_type="image/png")

    metadata = {
        "table": table_ref,
        "selector": selector,
        "source_file_id": file_id,
        "kind": chosen_kind,
        "x": chosen_x,
        "y": chosen_y,
        "generated_at": datetime.utcnow().isoformat(),
    }

    await ctx.db.execute(
        text(
            """
            INSERT INTO ekchat.plot_artifacts (
                id,
                chat_id,
                tenant_id,
                user_id,
                name,
                blob_path,
                metadata,
                created_at
            )
            VALUES (
                :id,
                :chat_id,
                :tenant_id,
                :user_id,
                :name,
                :blob_path,
                CAST(:metadata AS JSONB),
                NOW()
            )
            """
        ),
        {
            "id": artifact_id,
            "chat_id": resolved_chat_id,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
            "name": artifact_name,
            "blob_path": blob_path,
            "metadata": json.dumps(metadata),
        },
    )

    return {
        "ok": True,
        "plot": {
            "id": artifact_id,
            "name": artifact_name,
            "kind": chosen_kind,
            "x": chosen_x,
            "y": chosen_y,
            "url": f"/api/ekchat/v1/data/plots/{resolved_chat_id}/{artifact_name}",
        },
    }


@router.get("/data/plots/{chat_id}/{name}")
async def data_plot_download(
    chat_id: str,
    name: str,
    ctx: RequestContext = Depends(require_feature("ekchat_rag_enabled", default=False)),
):
    await _load_chat(ctx, chat_id)

    result = await ctx.db.execute(
        text(
            """
            SELECT blob_path
            FROM ekchat.plot_artifacts
            WHERE chat_id = :chat_id
              AND name = :name
              AND tenant_id = :tenant_id
              AND user_id = :user_id
            ORDER BY created_at DESC
            LIMIT 1
            """
        ),
        {
            "chat_id": chat_id,
            "name": name,
            "tenant_id": ctx.user.tenant_id,
            "user_id": ctx.user.user_id,
        },
    )

    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Plot artifact not found")

    storage = get_storage()
    image = await storage.download_bytes(row["blob_path"])

    return Response(
        content=image,
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="{Path(name).name}"'},
    )
